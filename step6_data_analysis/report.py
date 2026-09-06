#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
报告生成器
- Markdown 日报
- Excel 多 sheet 报告
"""
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill


# ============================================================
# Markdown 日报
# ============================================================

def generate_markdown(results, output_path):
    """生成 Markdown 格式日报"""
    now = datetime.now().strftime('%Y-%m-%d %H:%M')
    lines = []
    lines.append('# 微博数据分析日报')
    lines.append('')
    lines.append(f'> 生成时间：{now}')
    lines.append('')

    # 一、数据概览
    stats = results['stats']
    lines.append('## 一、数据概览')
    lines.append('')
    lines.append('| 指标 | 数量 |')
    lines.append('|---|---|')
    lines.append(f'| 微博帖子 | {stats["posts"]:,} |')
    lines.append(f'| 微博评论 | {stats["comments"]:,} |')
    lines.append(f'| 用户数 | {stats["users"]:,} |')
    lines.append('')

    # 二、热点排行
    lines.append('## 二、热点微博排行 TOP20')
    lines.append('')
    lines.append('| 排名 | 作者 | 内容摘要 | 点赞 | 评论 | 转发 | 热点指数 |')
    lines.append('|---|---|---|---|---|---|---|')
    for i, p in enumerate(results['hotspot'], 1):
        content = (p.get('content') or '')[:30].replace('\n', ' ').replace('|', '/')
        lines.append(
            f'| {i} | {p.get("username","")} | {content} | '
            f'{p.get("like_count",0):,} | {p.get("comment_count",0):,} | '
            f'{p.get("repost_count",0):,} | {p.get("hotspot_score",0)} |'
        )
    lines.append('')

    # 三、关键词趋势
    lines.append('## 三、关键词趋势')
    lines.append('')
    lines.append('| 关键词 | 帖子提及 | 评论提及 | 总提及 |')
    lines.append('|---|---|---|---|')
    for kw in results['keywords']:
        lines.append(f'| {kw["keyword"]} | {kw["post_mentions"]} | {kw["comment_mentions"]} | {kw["total_mentions"]} |')
    lines.append('')

    # 四、情感分析
    sa = results['sentiment']
    lines.append('## 四、评论情感分析')
    lines.append('')
    lines.append(f'- 分析评论数：{sa["total_analyzed"]:,}（采样上限 {sa["sample_size"]}）')
    lines.append(f'- 正面：{sa["positive_count"]:,}（{sa["positive_ratio"]}%）')
    lines.append(f'- 中性：{sa["neutral_count"]:,}（{sa["neutral_ratio"]}%）')
    lines.append(f'- 负面：{sa["negative_count"]:,}（{sa["negative_ratio"]}%）')
    lines.append('')
    lines.append('### 高频负面观点')
    lines.append('')
    lines.append('| 排名 | 关键词 | 出现次数 |')
    lines.append('|---|---|---|')
    for i, (word, cnt) in enumerate(sa['top_negative_viewpoints'], 1):
        lines.append(f'| {i} | {word} | {cnt} |')
    lines.append('')
    if sa['top_negative_comments']:
        lines.append('### 高赞负面评论')
        lines.append('')
        for i, c in enumerate(sa['top_negative_comments'][:5], 1):
            lines.append(f'{i}. **{c["username"]}**（👍{c["like_count"]}）：{c["text"]}')
        lines.append('')

    # 五、用户影响力
    ui = results['users']
    lines.append('## 五、用户影响力')
    lines.append('')
    lines.append('### 高粉丝用户 TOP10')
    lines.append('')
    lines.append('| 排名 | 用户 | 粉丝数 | 关注数 | 微博数 |')
    lines.append('|---|---|---|---|---|')
    for i, u in enumerate(ui['top_followers'][:10], 1):
        lines.append(
            f'| {i} | {u.get("username","")} | {u.get("followers_count",0):,} | '
            f'{u.get("following_count",0):,} | {u.get("weibo_count",0):,} |'
        )
    lines.append('')
    lines.append('### 高互动用户 TOP10')
    lines.append('')
    lines.append('| 排名 | 用户 | 帖子数 | 总点赞 | 总评论 | 总转发 | 总互动 |')
    lines.append('|---|---|---|---|---|---|---|')
    for i, e in enumerate(ui['top_engagement'][:10], 1):
        lines.append(
            f'| {i} | {e["username"]} | {e["post_count"]} | '
            f'{e["total_likes"]:,} | {e["total_comments"]:,} | '
            f'{e["total_reposts"]:,} | {e["total_engagement"]:,} |'
        )
    lines.append('')

    with open(output_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))
    return output_path


# ============================================================
# Excel 多 sheet 报告
# ============================================================

def _auto_width(ws):
    """自动调整列宽"""
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                val = str(cell.value) if cell.value is not None else ''
                # 中文字符算 2 宽度
                width = sum(2 if ord(c) > 127 else 1 for c in val)
                if width > max_len:
                    max_len = width
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)


def _style_header(ws):
    """表头加粗+底色"""
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')


def generate_excel(results, output_path):
    """生成 Excel 多 sheet 报告"""
    wb = Workbook()

    # Sheet 1: 数据概览
    ws = wb.active
    ws.title = '数据概览'
    stats = results['stats']
    ws.append(['指标', '数量'])
    ws.append(['微博帖子', stats['posts']])
    ws.append(['微博评论', stats['comments']])
    ws.append(['用户数', stats['users']])
    _style_header(ws)
    _auto_width(ws)

    # Sheet 2: 热点排行
    ws2 = wb.create_sheet('热点排行')
    ws2.append(['排名', 'weibo_id', '作者', '内容摘要', '点赞', '评论', '转发', '发布时间', '热点指数'])
    for i, p in enumerate(results['hotspot'], 1):
        ws2.append([
            i, p.get('weibo_id', ''), p.get('username', ''),
            (p.get('content') or '')[:100],
            p.get('like_count', 0), p.get('comment_count', 0),
            p.get('repost_count', 0), str(p.get('publish_time', '')),
            p.get('hotspot_score', 0),
        ])
    _style_header(ws2)
    _auto_width(ws2)

    # Sheet 3: 关键词趋势
    ws3 = wb.create_sheet('关键词趋势')
    ws3.append(['关键词', '帖子提及', '评论提及', '总提及'])
    for kw in results['keywords']:
        ws3.append([kw['keyword'], kw['post_mentions'], kw['comment_mentions'], kw['total_mentions']])
    _style_header(ws3)
    _auto_width(ws3)

    # Sheet 4: 情感分析
    sa = results['sentiment']
    ws4 = wb.create_sheet('情感分析')
    ws4.append(['指标', '数值'])
    ws4.append(['分析评论数', sa['total_analyzed']])
    ws4.append(['采样上限', sa['sample_size']])
    ws4.append(['正面数', sa['positive_count']])
    ws4.append(['正面比例%', sa['positive_ratio']])
    ws4.append(['中性数', sa['neutral_count']])
    ws4.append(['中性比例%', sa['neutral_ratio']])
    ws4.append(['负面数', sa['negative_count']])
    ws4.append(['负面比例%', sa['negative_ratio']])
    ws4.append([])
    ws4.append(['排名', '负面关键词', '出现次数'])
    for i, (word, cnt) in enumerate(sa['top_negative_viewpoints'], 1):
        ws4.append([i, word, cnt])
    ws4.append([])
    ws4.append(['高赞负面评论', '用户', '点赞'])
    for c in sa['top_negative_comments'][:10]:
        ws4.append([c['text'][:100], c['username'], c['like_count']])
    _style_header(ws4)
    _auto_width(ws4)

    # Sheet 5: 高粉丝用户
    ui = results['users']
    ws5 = wb.create_sheet('高粉丝用户')
    ws5.append(['排名', 'user_id', '用户名', '粉丝数', '关注数', '微博数', '认证', '简介'])
    for i, u in enumerate(ui['top_followers'], 1):
        ws5.append([
            i, u.get('user_id', ''), u.get('username', ''),
            u.get('followers_count', 0), u.get('following_count', 0),
            u.get('weibo_count', 0), u.get('verified', 0),
            (u.get('description') or '')[:80],
        ])
    _style_header(ws5)
    _auto_width(ws5)

    # Sheet 6: 高互动用户
    ws6 = wb.create_sheet('高互动用户')
    ws6.append(['排名', 'user_id', '用户名', '帖子数', '总点赞', '总评论', '总转发', '总互动'])
    for i, e in enumerate(ui['top_engagement'], 1):
        ws6.append([
            i, e['user_id'], e['username'], e['post_count'],
            e['total_likes'], e['total_comments'], e['total_reposts'], e['total_engagement'],
        ])
    _style_header(ws6)
    _auto_width(ws6)

    wb.save(output_path)
    return output_path
